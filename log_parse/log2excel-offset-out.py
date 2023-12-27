
import sys
import datetime

if len(sys.argv) != 2:      # 아규먼트 2개만 
    print("""
usage:
    python log2excel.py [1-9]
    """)
    sys.exit()

tms_num: str = sys.argv[1]
print("TMS Number: ", tms_num)

import re
from openpyxl import Workbook

log_file = open("20200728-1.log", 'r')      # 로그 파일명 (수정 필요)

pre_str: str = 'DEV[ '+ tms_num +'] receive temperature in '

out_wb = Workbook()
out_ws = out_wb.active


out_ws.cell(row=1, column=1).value = 'Time'
out_ws.cell(row=1, column=2).value = 'TMS_' + tms_num
ro: int = 2

time_s = datetime.datetime(2020, 7, 29, 9, 10, 00) # 로그 시작 시간 (수정 필요)
time_e = datetime.datetime(2020, 7, 29, 14, 50, 00) # 로그 종료 시간 (수정 필요)
print(time_s)

while True:
    line: str = log_file.readline()
    if not line: 
        break
    exist: int = line.find(pre_str)
    if exist != -1:
#        print(line, end="")
        time: str = re.search(r'\d{1,4}-\d{1,2}-\d{1,2} \d\d:\d\d:\d\d', line).group(0)
#         print('time: ' + time)
        time_cur = datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
        if time_cur < time_s:
            continue;
        time_diff = ((time_cur-time_s).seconds) // 10  # 10초 간격 off set 적용
        if time_cur > time_e :
            break;

        temp: str = re.search(r'out\([+-]?\d{1,4}[.]\d{1,4}\/', line).group(0)
#        print('temp: ' + temp)
        celius: str = re.search(r'[+-]?\d{1,3}[.]\d{1,2}', temp).group(0)
#        print('celius: ' + celius)

        out_ws.cell(row=ro, column=1).value = time
        out_ws.cell(row=ro, column=2).value = float(celius)
        ro = time_diff + 2

out_wb.save('TMS_' + tms_num + '.xlsx')

log_file.close()

