
import sys
import datetime
import openpyxl


import re
from openpyxl import Workbook

in_wb = openpyxl.load_workbook('./계측기-데이터.xlsx')
in_ws = in_wb.active

out_wb = Workbook()
out_ws = out_wb.active

time_s = datetime.datetime(2020, 7, 29, 9, 10, 00) # 로그 시작 시간 (수정 필요)
time_e = datetime.datetime(2020, 7, 29, 14, 50, 00) # 로그 종료 시간 (수정 필요)

ro = 1
out_ws.cell(row=ro, column=1).value = 'time'
out_ws.cell(row=ro, column=2).value = 'dev1'
out_ws.cell(row=ro, column=3).value = 'dev2'
out_ws.cell(row=ro, column=4).value = 'dev3'
out_ws.cell(row=ro, column=5).value = 'dev4'
out_ws.cell(row=ro, column=6).value = 'dev5'

for r in in_ws.rows:
    row_index = r[0].row
    time = r[0].value
    m1 = r[1].value
    m2 = r[2].value
    m3 = r[3].value
    m4 = r[4].value
    m5 = r[5].value

    time_cur = datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S")

    if time_cur < time_s:
        continue;
    time_diff = ((time_cur-time_s).seconds) // 10  # 10초 간격 off set 적용
    if time_cur > time_e :
        break;

    ro = time_diff + 2

    out_ws.cell(row=ro, column=1).value = time
    out_ws.cell(row=ro, column=2).value = float(m1)
    out_ws.cell(row=ro, column=3).value = float(m2)
    out_ws.cell(row=ro, column=4).value = float(m3)
    out_ws.cell(row=ro, column=5).value = float(m4)
    out_ws.cell(row=ro, column=6).value = float(m5)

out_wb.save('Meter_.xlsx')







