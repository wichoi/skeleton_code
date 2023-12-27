
import sys

if len(sys.argv) != 2:      # 아규먼트 2개만 
    print("""
usage:
    python log2excel.py [1-9]
    """)
    sys.exit()

tms_num: str = sys.argv[1]
print("TMS Number: ", tms_num)

import re
from openpyxl import Workbook

log_file = open("20200624-1.log", 'r')

pre_str: str = 'DEV[ '+ tms_num +'] receive temperature in '

out_wb = Workbook()
out_ws = out_wb.active


out_ws.cell(row=1, column=1).value = 'Time'
out_ws.cell(row=1, column=2).value = 'TMS_' + tms_num
ro: int = 2

while True:
    line: str = log_file.readline()
    if not line: 
        break
    exist: int = line.find(pre_str)
    if exist != -1:
#        print(line, end="")
        time: str = re.search(r'\d\d:\d\d:\d\d', line).group(0)
#         print('time: ' + time)

        temp: str = re.search(r'\([+-]?\d{1,3}[.]\d{1,2}\/.{1,3}\)', line).group(0)
#        print('temp: ' + temp)
        celius: str = re.search(r'[+-]?\d{1,3}[.]\d{1,2}', temp).group(0)
#        print('celius: ' + celius)

        out_ws.cell(row=ro, column=1).value = time
        out_ws.cell(row=ro, column=2).value = float(celius)
        ro += 6

out_wb.save('TMS_' + tms_num + '.xlsx')

log_file.close()





