
import sys

if len(sys.argv) != 2:      # 아규먼트 2개만 
    print("""
usage:
    python log2excel.py [1-9]
    """)
    sys.exit()

tms_num: str = sys.argv[1]
print("TMS Number: ", tms_num)

import re
from openpyxl import Workbook

log_file = open("20200811-05.log", 'r')

pre_str: str = 'DEV[ '+ tms_num +'] receive '

out_wb = Workbook()
out_ws = out_wb.active


out_ws.cell(row=1, column=1).value = 'Time'
out_ws.cell(row=1, column=2).value = 'temperature'
out_ws.cell(row=1, column=3).value = 'bat'
ro: int = 2

while True:
    line: str = log_file.readline()
    if not line: 
        break
    exist: int = line.find(pre_str)
    if exist != -1:
#        print(line, end="")
        time: str = re.search(r'\d\d:\d\d:\d\d', line).group(0)
#         print('time: ' + time)

        temp: str = re.search(r'in \([+-]?\d{1,4}[.]\d{1,4}\/', line).group(0)
        temperature: str = re.search(r'[+-]?\d{1,4}[.]\d{1,4}', temp).group(0)

        temp: str = re.search(r'bat\( \d\d\d\d\)', line).group(0)
        bat: str = re.search(r'\d\d\d\d', temp).group(0)


        out_ws.cell(row=ro, column=1).value = time
        out_ws.cell(row=ro, column=2).value = temperature
        out_ws.cell(row=ro, column=3).value = bat
        ro += 1

out_wb.save('TMS_' + tms_num + '.xlsx')

log_file.close()





